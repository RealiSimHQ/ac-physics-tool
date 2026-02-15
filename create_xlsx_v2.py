#!/usr/bin/env python3
"""Create XLSX matching Ryan's spreadsheet format with Drive share links and deduplication."""
import csv, json, subprocess, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Get file IDs from Google Drive for share links
def get_drive_file_ids(remote, path):
    """Get file IDs for generating share links."""
    result = subprocess.run(
        ['rclone', 'lsjson', f'{remote}:{path}', '-R', '--no-modtime'],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print(f"  ⚠️ Could not get IDs for {path}: {result.stderr[:100]}")
        return {}
    
    files = json.loads(result.stdout)
    id_map = {}
    for f in files:
        if not f.get('IsDir', False) and f.get('ID'):
            id_map[f['Path']] = f['ID']
    return id_map

def make_drive_link(file_id):
    """Create a Google Drive download/view link."""
    return f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"

def clean_name(filename):
    """Extract clean track/car name from filename."""
    name = filename
    for ext in ['.7z', '.zip', '.rar']:
        name = name.replace(ext, '')
    # Remove common prefixes
    prefixes = ['X10DD.', 'X10DD-', 'X10DD.Battle-', 'X10DD.Grassroots-', 'X10DD.Street-', 
                'X10DD.Street+', 'X10DD.ProSpec-', 'X10DD.ProSpec+', 'X10DD.Muscle-',
                'X10DD.Arch+', 'X10DD.Arch-', 'X10DD.FnF-',
                'TANDEM-', 'TAP-', 'TAProot-', 'JDAM-', 'JDAM_', 'JDAM ',
                'LVL-', 'LVL-240_55Deg-', 'LVL-250_50Deg-', 'LVL420-',
                'USD-', 'IUSD-', 'PDP_', '-Ai.5O-', '-AiO.Street-',
                'LVLRX_', 'LVLGR8_', 'LAWD-',
                'adl_proam_', 'adl-proam-',
                '_RealiSimHQ_', 'RealiSim HQ Layout_']
    for p in prefixes:
        if name.startswith(p):
            name = name[len(p):]
            break
    # Clean up underscores and extra spaces
    name = name.replace('_', ' ').strip()
    # Remove leading/trailing special chars
    name = name.strip('-_ .')
    return name

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
link_font = Font(color='1155CC', underline='single')
section_font = Font(bold=True, size=13)
section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
section_text = Font(bold=True, color='FFFFFF', size=13)

# Make/Model splitter
MAKE_PATTERNS = [
    ('BMW', ['BMW']),
    ('Chevrolet', ['Chevy', 'Chevrolet', 'Corvette', 'Camaro', 'Chevelle', 'Nova', 'Impala']),
    ('Ford', ['Ford']),
    ('Toyota', ['Toyota', 'Scion']),
    ('Nissan', ['Nissan', 'Infiniti', 'Infinity']),
    ('Mazda', ['Mazda']),
    ('Honda', ['Honda', 'Acura']),
    ('Subaru', ['Subaru']),
    ('Mitsubishi', ['Mitsubishi']),
    ('Dodge', ['Dodge']),
    ('Plymouth', ['Plymouth']),
    ('Pontiac', ['Pontiac']),
    ('Buick', ['Buick']),
    ('GMC', ['GMC']),
    ('Cadillac', ['Cadillac']),
    ('Mercury', ['Mercury']),
    ('Lexus', ['Lexus']),
    ('Porsche', ['Porsche']),
    ('Volkswagen', ['Volkswagen', 'VW']),
    ('Hyundai', ['Hyundai']),
    ('Lancia', ['Lancia', 'Delta']),
    ('Peugeot', ['Peugeot']),
    ('MG', ['MG']),
    ('Suzuki', ['Suzuki']),
    ('Kawasaki', ['Kawasaki']),
    ('Yamaha', ['Yamaha']),
    ('Daihatsu', ['Daihatsu']),
    ('Isuzu', ['Isuzu']),
    ('Audi', ['Audi']),
    ('Mercedes', ['Mercedes']),
    ('Tesla', ['Teslide', 'Tesla']),
]

def split_make_model(name):
    """Split a car name into (Make, Model)."""
    name_lower = name.lower()
    for make, patterns in MAKE_PATTERNS:
        for pat in patterns:
            pat_lower = pat.lower()
            if pat_lower in name_lower:
                # Find where the pattern is and take the rest as model
                idx = name_lower.find(pat_lower)
                # Model is everything after the make word
                remainder = name[idx + len(pat):].strip(' -_')
                if not remainder:
                    remainder = name[:idx].strip(' -_') or name
                return make, remainder
    # Fallback: first word = make, rest = model
    parts = name.split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return name, ''

print("📂 Getting file IDs from Google Drive...")

# Get IDs for share links
print("  Scanning Tracks/Released...")
track_ids_released = get_drive_file_ids('gdrive', 'Tracks/Released')
print(f"    {len(track_ids_released)} files")

print("  Scanning Assetto Corsa/Tracks...")
track_ids_ac = get_drive_file_ids('gdrive', 'Assetto Corsa/Tracks')
print(f"    {len(track_ids_ac)} files")

print("  Scanning Assetto Corsa/Cars...")
car_ids = get_drive_file_ids('gdrive', 'Assetto Corsa/Cars')
print(f"    {len(car_ids)} files")

# Build track data with dedup
print("\n📊 Building deduplicated catalog...")

# Combine all tracks, dedup by cleaned name
all_tracks = {}  # clean_name -> {name, category, links: [(path, id, subcategory)]}

def add_tracks(id_map, base_path):
    for path, fid in id_map.items():
        filename = path.split('/')[-1]
        name = clean_name(filename)
        name_key = name.lower().replace(' ', '').replace('-', '').replace('_', '')
        
        # Get category from path
        parts = path.split('/')
        category = parts[0] if len(parts) > 1 else 'Other'
        subcategory = '/'.join(parts[1:-1]) if len(parts) > 2 else ''
        
        if name_key not in all_tracks:
            all_tracks[name_key] = {
                'name': name,
                'category': category,
                'links': [],
            }
        all_tracks[name_key]['links'].append({
            'filename': filename,
            'path': f"{base_path}/{path}",
            'id': fid,
            'subcategory': f"{category}/{subcategory}" if subcategory else category,
        })

add_tracks(track_ids_released, 'Tracks/Released')
add_tracks(track_ids_ac, 'Assetto Corsa/Tracks')

# Build car data grouped by series
car_series = defaultdict(list)  # series -> [(name, filename, id, subcategory)]
for path, fid in car_ids.items():
    parts = path.split('/')
    series = parts[0] if len(parts) > 1 else 'Other'
    filename = parts[-1]
    name = clean_name(filename)
    subcategory = '/'.join(parts[1:-1]) if len(parts) > 2 else ''
    car_series[series].append({
        'name': name,
        'filename': filename,
        'id': fid,
        'subcategory': subcategory,
    })

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# === TRACKS TAB ===
ws_tracks = wb.create_sheet(title='Tracks')
headers = ['Track Name', 'Drive Link', 'Category', 'Other Versions']
for c, h in enumerate(headers, 1):
    cell = ws_tracks.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

row = 2
for name_key in sorted(all_tracks.keys()):
    track = all_tracks[name_key]
    links = track['links']
    
    # Primary link (first one)
    primary = links[0]
    
    cell_name = ws_tracks.cell(row=row, column=1, value=track['name'])
    
    # Drive link as hyperlink
    url = make_drive_link(primary['id'])
    cell_link = ws_tracks.cell(row=row, column=2, value=primary['filename'])
    cell_link.hyperlink = url
    cell_link.font = link_font
    
    ws_tracks.cell(row=row, column=3, value=primary['subcategory'])
    
    # Other versions
    if len(links) > 1:
        other_links = []
        for l in links[1:]:
            other_links.append(l['filename'])
        ws_tracks.cell(row=row, column=4, value=', '.join(other_links))
    
    if row % 2 == 0:
        for c in range(1, 5):
            ws_tracks.cell(row=row, column=c).fill = alt_fill
    
    row += 1

# Column widths
ws_tracks.column_dimensions['A'].width = 35
ws_tracks.column_dimensions['B'].width = 45
ws_tracks.column_dimensions['C'].width = 30
ws_tracks.column_dimensions['D'].width = 50
ws_tracks.freeze_panes = 'A2'
ws_tracks.auto_filter.ref = f"A1:D{row-1}"

print(f"  ✅ Tracks: {len(all_tracks)} unique tracks (deduped from {sum(len(t['links']) for t in all_tracks.values())} files)")

# === CAR TABS (one per series) ===
series_order = ['X10DD', 'L.E.V.E.L.', 'T.A.N.D.E.M.', 'ADL-RealiSim HQ', 'TAP', 
                'JDAM', 'RallyCross', 'Initial D', 'U$D v2', 'Traffic', 'Police',
                'Zipped Carpacks', 'RealiSimHQ Parts', 'Initial U$D']

for series in series_order:
    if series not in car_series:
        continue
    cars = car_series[series]
    
    # Clean tab name (max 31 chars, no special chars)
    tab_name = series[:31].replace('/', '-').replace('$', 'S')
    
    ws = wb.create_sheet(title=tab_name)
    car_headers = ['Car Name', 'Make', 'Model', 'Drive Link', 'Subcategory']
    for c, h in enumerate(car_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for r, car in enumerate(sorted(cars, key=lambda x: x['name']), 2):
        make, model = split_make_model(car['name'])
        ws.cell(row=r, column=1, value=car['name'])
        ws.cell(row=r, column=2, value=make)
        ws.cell(row=r, column=3, value=model)
        
        url = make_drive_link(car['id'])
        cell_link = ws.cell(row=r, column=4, value=car['filename'])
        cell_link.hyperlink = url
        cell_link.font = link_font
        
        ws.cell(row=r, column=5, value=car['subcategory'])
        
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = alt_fill
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 25
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:E{len(cars)+1}"
    
    print(f"  ✅ {series}: {len(cars)} cars")

# Handle any series not in our order
for series in car_series:
    if series not in series_order and series:
        cars = car_series[series]
        tab_name = series[:31].replace('/', '-').replace('$', 'S')
        ws = wb.create_sheet(title=tab_name)
        car_headers = ['Car Name', 'Make', 'Model', 'Drive Link', 'Subcategory']
        for c, h in enumerate(car_headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r, car in enumerate(sorted(cars, key=lambda x: x['name']), 2):
            make, model = split_make_model(car['name'])
            ws.cell(row=r, column=1, value=car['name'])
            ws.cell(row=r, column=2, value=make)
            ws.cell(row=r, column=3, value=model)
            url = make_drive_link(car['id'])
            cell_link = ws.cell(row=r, column=4, value=car['filename'])
            cell_link.hyperlink = url
            cell_link.font = link_font
            ws.cell(row=r, column=5, value=car['subcategory'])
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 25
        print(f"  ✅ {series}: {len(cars)} cars")

# Save
output = '/tmp/RealiSimHQ_AC_Content_Catalog.xlsx'
wb.save(output)
print(f"\n📁 Saved: {output}")
print(f"📊 {len(all_tracks)} tracks + {sum(len(v) for v in car_series.values())} cars")
