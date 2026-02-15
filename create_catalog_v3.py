#!/usr/bin/env python3
"""Create a polished RealiSimHQ AC Content Catalog spreadsheet with cover sheet."""
import csv, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CATALOG_DIR = os.path.join(os.path.dirname(__file__), 'catalog')
GDRIVE_BASE = "https://drive.google.com/file/d/{}/view?usp=drive_link"

# Color palette
CYAN = "00FFFF"
DARK_BG = "1a1a2e"
DARKER_BG = "16213e"
HEADER_BG = "0f3460"
ACCENT = "e94560"
GOLD = "FFD700"
WHITE = "FFFFFF"
LIGHT_GRAY = "CCCCCC"
MED_GRAY = "888888"
ROW_EVEN = "1e2a4a"
ROW_ODD = "16213e"

# Styles
header_font = Font(name="Calibri", size=12, bold=True, color=CYAN)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
title_font = Font(name="Calibri", size=22, bold=True, color=CYAN)
subtitle_font = Font(name="Calibri", size=14, color=GOLD)
body_font = Font(name="Calibri", size=11, color=WHITE)
link_font = Font(name="Calibri", size=11, color=CYAN, underline="single")
accent_font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
stat_font = Font(name="Calibri", size=16, bold=True, color=GOLD)
stat_label_font = Font(name="Calibri", size=11, color=LIGHT_GRAY)
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
darker_fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")
even_fill = PatternFill(start_color=ROW_EVEN, end_color=ROW_EVEN, fill_type="solid")
odd_fill = PatternFill(start_color=ROW_ODD, end_color=ROW_ODD, fill_type="solid")
thin_border = Border(
    bottom=Side(style="thin", color="333355")
)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_sheet_dark(ws, num_cols=6):
    """Apply dark background to entire visible area."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 50, min_col=1, max_col=num_cols + 5):
        for cell in row:
            cell.fill = dark_fill


def parse_car_name(raw_name):
    """Try to extract Make and Model from car name."""
    name = raw_name
    # Remove common prefixes
    for prefix in ['X10DD.Battle-', 'X10DD.Drift-', 'X10DD.Track-', 'X10DD.Street-', 'X10DD.Rally-',
                   'X10DD-', 'TAP-', 'TAProot-', 'USD-', 'LVL-', 'LVLGR8_', 'LVLRX_', 'LVLRX-',
                   'adl-proam-', 'adl_proam_', 'JDAM-', 'JDAM_', 'tdm_', 'tdm-']:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    
    # Common make mappings
    makes = {
        'BMW': ['BMW', 'bmw'],
        'Nissan': ['Nissan', 'nissan'],
        'Toyota': ['Toyota', 'toyota', 'AE86', 'Supra'],
        'Honda': ['Honda', 'honda', 'Civic', 'S2000'],
        'Mazda': ['Mazda', 'mazda', 'RX-7', 'RX7', 'Miata', 'MX-5'],
        'Chevrolet': ['Chevy', 'chevy', 'Chevrolet', 'Corvette', 'Chevelle', 'C6'],
        'Ford': ['Ford', 'ford', 'Mustang', 'RS200'],
        'Subaru': ['Subaru', 'subaru', 'WRX', 'Impreza'],
        'Mitsubishi': ['Mitsubishi', 'mitsubishi', 'Evo', 'Lancer'],
        'Dodge': ['Dodge', 'dodge', 'Charger', 'Challenger'],
        'Suzuki': ['Suzuki', 'suzuki', 'Cappuccino', 'Hayabusa', 'Swift'],
        'Kawasaki': ['Kawasaki', 'kawasaki', 'ZX'],
        'Yamaha': ['Yamaha', 'yamaha'],
        'Lancia': ['Lancia', 'lancia', 'Delta'],
        'Peugeot': ['Peugeot', 'peugeot', '205'],
        'Porsche': ['Porsche', 'porsche', '959'],
        'MG': ['MG', 'Metro'],
        'Volkswagen': ['VW', 'Volkswagen', 'Golf'],
    }
    
    detected_make = "Other"
    for make, keywords in makes.items():
        for kw in keywords:
            if kw.lower() in name.lower():
                detected_make = make
                break
        if detected_make != "Other":
            break
    
    # Clean up model name
    model = name.replace('_', ' ').replace('-', ' ').strip()
    
    return detected_make, model


def load_cars():
    cars = []
    with open(os.path.join(CATALOG_DIR, 'cars.csv')) as f:
        for row in csv.DictReader(f):
            if not row['Category']:
                continue
            make, model = parse_car_name(row['Name'])
            cars.append({
                'name': row['Name'],
                'make': make,
                'model': model,
                'pack': row['Category'],
                'sub': row.get('Subcategory', ''),
                'file': row['Filename'],
                'size_mb': float(row['Size (MB)']) if row['Size (MB)'] else 0,
                'path': row['Full Path'],
            })
    return cars


def load_tracks():
    tracks = []
    seen = set()
    for csv_file in ['tracks_ac_folder.csv', 'tracks_released.csv']:
        path = os.path.join(CATALOG_DIR, csv_file)
        if not os.path.exists(path):
            continue
        with open(path) as f:
            for row in csv.DictReader(f):
                key = row['Name'].lower().strip()
                if key in seen:
                    continue
                seen.add(key)
                tracks.append({
                    'name': row['Name'],
                    'category': row.get('Category', ''),
                    'sub': row.get('Subcategory', ''),
                    'file': row['Filename'],
                    'size_mb': float(row['Size (MB)']) if row['Size (MB)'] else 0,
                    'path': row['Full Path'],
                })
    return tracks


def create_dashboard(wb, cars, tracks):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = ACCENT
    
    # Column widths
    for col, width in [(1, 3), (2, 25), (3, 25), (4, 25), (5, 20), (6, 15), (7, 3)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Dark background
    for row in range(1, 45):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = dark_fill
    
    # Title
    ws.merge_cells('B2:F2')
    c = ws.cell(row=2, column=2, value="RealiSimHQ Content Catalog")
    c.font = title_font
    c.fill = dark_fill
    c.alignment = center
    
    ws.merge_cells('B3:F3')
    c = ws.cell(row=3, column=2, value="Assetto Corsa  ·  Extended Physics  ·  COSMIC Suspension")
    c.font = Font(name="Calibri", size=12, color=LIGHT_GRAY)
    c.fill = dark_fill
    c.alignment = center
    
    # Stats row
    row = 5
    stats = [
        ("Cars", len(cars)),
        ("Tracks", len(tracks)),
        ("Car Packs", len(set(c['pack'] for c in cars))),
        ("Total Size", f"{sum(c['size_mb'] for c in cars) + sum(t['size_mb'] for t in tracks):.1f} GB"),
    ]
    for i, (label, value) in enumerate(stats):
        col = 2 + i
        c = ws.cell(row=row, column=col, value=str(value))
        c.font = stat_font
        c.fill = darker_fill
        c.alignment = center
        c = ws.cell(row=row + 1, column=col, value=label)
        c.font = stat_label_font
        c.fill = darker_fill
        c.alignment = center
    
    # Filter section
    row = 8
    ws.merge_cells(f'B{row}:F{row}')
    c = ws.cell(row=row, column=2, value="━━━━━  QUICK FILTER  ━━━━━")
    c.font = Font(name="Calibri", size=14, bold=True, color=CYAN)
    c.fill = dark_fill
    c.alignment = center
    
    # Make dropdown
    row = 10
    makes = sorted(set(c['make'] for c in cars))
    packs = sorted(set(c['pack'] for c in cars))
    
    ws.cell(row=row, column=2, value="Filter by Make:").font = Font(name="Calibri", size=11, bold=True, color=GOLD)
    ws.cell(row=row, column=2).fill = dark_fill
    make_cell = ws.cell(row=row, column=3, value="All")
    make_cell.font = body_font
    make_cell.fill = darker_fill
    make_cell.alignment = center
    make_dv = DataValidation(type="list", formula1='"All,' + ','.join(makes) + '"', allow_blank=True)
    make_dv.prompt = "Select a car make"
    ws.add_data_validation(make_dv)
    make_dv.add(make_cell)
    
    # Pack dropdown
    row = 11
    ws.cell(row=row, column=2, value="Filter by Pack:").font = Font(name="Calibri", size=11, bold=True, color=GOLD)
    ws.cell(row=row, column=2).fill = dark_fill
    pack_cell = ws.cell(row=row, column=3, value="All")
    pack_cell.font = body_font
    pack_cell.fill = darker_fill
    pack_cell.alignment = center
    pack_dv = DataValidation(type="list", formula1='"All,' + ','.join(packs) + '"', allow_blank=True)
    pack_dv.prompt = "Select a car pack"
    ws.add_data_validation(pack_dv)
    pack_dv.add(pack_cell)
    
    # Instructions
    row = 13
    ws.merge_cells(f'B{row}:F{row}')
    c = ws.cell(row=row, column=2, value="Use dropdowns above, then check the Cars tab — use Excel/Sheets Filter to match.")
    c.font = Font(name="Calibri", size=10, italic=True, color=MED_GRAY)
    c.fill = dark_fill
    
    # Quick links section
    row = 15
    ws.merge_cells(f'B{row}:F{row}')
    c = ws.cell(row=row, column=2, value="━━━━━  SHEET TABS  ━━━━━")
    c.font = Font(name="Calibri", size=14, bold=True, color=CYAN)
    c.fill = dark_fill
    c.alignment = center
    
    tabs = [
        ("🏎️  Cars", "Full car catalog with Make, Model, Pack, Size"),
        ("🏁  Tracks", "All released tracks with size and category"),
        ("📦  Car Packs", "Summary by car pack series"),
        ("🖥️  Servers", "Active server configurations"),
    ]
    for i, (name, desc) in enumerate(tabs):
        r = row + 2 + i
        c = ws.cell(row=r, column=2, value=name)
        c.font = Font(name="Calibri", size=12, bold=True, color=CYAN)
        c.fill = darker_fill
        c = ws.cell(row=r, column=3, value=desc)
        c.font = Font(name="Calibri", size=11, color=LIGHT_GRAY)
        c.fill = darker_fill
        ws.merge_cells(f'C{r}:F{r}')
    
    # Footer
    row = 24
    ws.merge_cells(f'B{row}:F{row}')
    c = ws.cell(row=row, column=2, value="RealiSimHQ  ·  Extended Physics  ·  Built by Paddy 🐘")
    c.font = Font(name="Calibri", size=10, color=MED_GRAY)
    c.fill = dark_fill
    c.alignment = center


def create_cars_sheet(wb, cars):
    ws = wb.create_sheet("Cars")
    ws.sheet_properties.tabColor = "00FFFF"
    
    # Column config
    cols = [
        ("Make", 16),
        ("Model", 35),
        ("Car Pack", 22),
        ("Physics Type", 14),
        ("Size (MB)", 12),
        ("Filename", 40),
    ]
    
    for i, (name, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    
    # Sort cars by make then model
    cars_sorted = sorted(cars, key=lambda x: (x['make'], x['model']))
    
    for idx, car in enumerate(cars_sorted):
        row = idx + 2
        fill = even_fill if idx % 2 == 0 else odd_fill
        
        values = [
            car['make'],
            car['model'],
            car['pack'],
            car['sub'] or '—',
            round(car['size_mb'], 1),
            car['file'],
        ]
        
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = body_font if col != 1 else Font(name="Calibri", size=11, bold=True, color=CYAN)
            c.fill = fill
            c.alignment = center if col in (4, 5) else left
            c.border = thin_border
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:F{len(cars_sorted) + 1}"
    
    # Freeze top row
    ws.freeze_panes = "A2"


def create_tracks_sheet(wb, tracks):
    ws = wb.create_sheet("Tracks")
    ws.sheet_properties.tabColor = "FFD700"
    
    cols = [
        ("Track Name", 45),
        ("Category", 22),
        ("Size (MB)", 12),
        ("Filename", 45),
    ]
    
    for i, (name, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    
    tracks_sorted = sorted(tracks, key=lambda x: x['name'])
    
    for idx, track in enumerate(tracks_sorted):
        row = idx + 2
        fill = even_fill if idx % 2 == 0 else odd_fill
        
        values = [
            track['name'],
            track['category'] or '—',
            round(track['size_mb'], 1),
            track['file'],
        ]
        
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = body_font
            c.fill = fill
            c.alignment = center if col == 3 else left
            c.border = thin_border
    
    ws.auto_filter.ref = f"A1:D{len(tracks_sorted) + 1}"
    ws.freeze_panes = "A2"


def create_packs_sheet(wb, cars):
    ws = wb.create_sheet("Car Packs")
    ws.sheet_properties.tabColor = "E94560"
    
    cols = [
        ("Car Pack", 25),
        ("# Cars", 10),
        ("Total Size (GB)", 16),
        ("Physics Types", 30),
        ("Sample Cars", 50),
    ]
    
    for i, (name, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    
    # Group by pack
    from collections import defaultdict
    packs = defaultdict(list)
    for car in cars:
        packs[car['pack']].append(car)
    
    pack_list = sorted(packs.items(), key=lambda x: -len(x[1]))
    
    for idx, (pack_name, pack_cars) in enumerate(pack_list):
        row = idx + 2
        fill = even_fill if idx % 2 == 0 else odd_fill
        total_gb = sum(c['size_mb'] for c in pack_cars) / 1024
        physics = sorted(set(c['sub'] for c in pack_cars if c['sub']))
        samples = ', '.join(c['model'][:25] for c in pack_cars[:4])
        
        values = [
            pack_name,
            len(pack_cars),
            round(total_gb, 2),
            ', '.join(physics) if physics else '—',
            samples + ('...' if len(pack_cars) > 4 else ''),
        ]
        
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = body_font if col != 1 else Font(name="Calibri", size=11, bold=True, color=CYAN)
            c.fill = fill
            c.alignment = center if col in (2, 3) else left
            c.border = thin_border
    
    ws.auto_filter.ref = f"A1:E{len(pack_list) + 1}"
    ws.freeze_panes = "A2"


def create_servers_sheet(wb):
    ws = wb.create_sheet("Servers")
    ws.sheet_properties.tabColor = "40E0D0"
    
    cols = [
        ("Server", 30),
        ("Theme", 25),
        ("HTTP Port", 12),
        ("TCP/UDP Port", 14),
        ("Cars", 8),
        ("Features", 40),
        ("Join Link", 55),
    ]
    
    for i, (name, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    
    servers = [
        ("01 TAProot Circuit", "Drift · Track Rotation", 9001, 10001, 6, "17 tracks, VotingPreset, Extended Physics", "https://acstuff.club/s/q:race/online/join?ip=152.53.83.105&httpPort=9001"),
        ("05 LA Canyons", "Open World · Traffic", 9004, 10004, 16, "AI Traffic, Open World, Extended Physics", "https://acstuff.club/s/q:race/online/join?ip=152.53.83.105&httpPort=9004"),
        ("06 Initial D", "Touge · Street Racing", 9006, 9606, 13, "X10DD Battle cars, Touge tracks", "https://acstuff.club/s/q:race/online/join?ip=152.53.83.105&httpPort=9006"),
        ("08 RallyX", "Rally · Mini Trucks", 9008, 9608, 11, "Group B + Hayabusa Mini Trucks, SurfaceFX", "https://acstuff.club/s/q:race/online/join?ip=152.53.83.105&httpPort=9008"),
    ]
    
    for idx, (name, theme, http, tcp, cars, features, link) in enumerate(servers):
        row = idx + 2
        fill = even_fill if idx % 2 == 0 else odd_fill
        
        for col, val in enumerate([name, theme, http, tcp, cars, features, link], 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 7:
                c.font = link_font
                c.hyperlink = val
            elif col == 1:
                c.font = Font(name="Calibri", size=11, bold=True, color=CYAN)
            else:
                c.font = body_font
            c.fill = fill
            c.alignment = center if col in (3, 4, 5) else left
            c.border = thin_border
    
    ws.freeze_panes = "A2"


def main():
    print("Loading catalog data...")
    cars = load_cars()
    tracks = load_tracks()
    
    print(f"  {len(cars)} cars, {len(tracks)} tracks")
    
    wb = Workbook()
    
    print("Creating Dashboard...")
    create_dashboard(wb, cars, tracks)
    
    print("Creating Cars sheet...")
    create_cars_sheet(wb, cars)
    
    print("Creating Tracks sheet...")
    create_tracks_sheet(wb, tracks)
    
    print("Creating Car Packs sheet...")
    create_packs_sheet(wb, cars)
    
    print("Creating Servers sheet...")
    create_servers_sheet(wb)
    
    out = os.path.join(os.path.dirname(__file__), 'RealiSimHQ_AC_Catalog_v3.xlsx')
    wb.save(out)
    print(f"\nSaved: {out}")
    print("Ready to upload to Google Drive!")


if __name__ == '__main__':
    main()
